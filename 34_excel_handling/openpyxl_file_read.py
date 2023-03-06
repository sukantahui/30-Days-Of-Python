# https://realpython.com/openpyxl-excel-spreadsheets-python/#getting-started-with-openpyxl
import os
os.system('cls')

from openpyxl import load_workbook

wb = load_workbook(filename="hello_world.xlsx")
sheet=wb.active
print(wb.sheetnames)
print(sheet.title)